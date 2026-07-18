#!/usr/bin/env python3
# scrape_enrich.py — stand-alone role-enrichment scraper (plain requests, MyCareersFuture).
#
# Pulls live SG job postings for SkillsFuture roles, matches each posting back to a
# role via TF-IDF (+ optional SSOC tie-breaker), and writes a cached enrichment
# artifact (data/role_enrichment.json) the pipeline later consumes read-only.
#
# No Oxylabs, no headless browser, no proxy. The MCF jobs API is a public,
# unauthenticated GET:  https://api.mycareersfuture.gov.sg/v2/jobs?search=&limit=&page=
#
# Run:
#   python scrape_enrich.py --endpoint-probe            # confirm endpoint, 1 request
#   python scrape_enrich.py --dry-run                    # parse+match on fixtures, 0 requests
#   python scrape_enrich.py --per-role 8 --track "Data and Artificial Intelligence"
#   python scrape_enrich.py --use-cache ...              # 0 live requests, replay cache
"""Stand-alone MyCareersFuture -> SkillsFuture role enrichment scraper."""
from __future__ import annotations

import argparse
import html as _html
import json
import os
import re
import sys
import time
from dataclasses import dataclass, field

import requests

try:
    from dotenv import load_dotenv
    load_dotenv()
except Exception:  # dotenv is optional
    pass

# --- paths (relative to this script's directory) ---
HERE = os.path.dirname(os.path.abspath(__file__))
SFW_XLSX = os.path.join(HERE, "jobsandskills-skillsfuture-skills-framework-dataset.xlsx")
SSOC_XLSX = os.path.join(HERE, "ssoc2024-classification-structure.xlsx")
DATA_DIR = os.path.join(HERE, "data")
CACHE_DIR = os.path.join(DATA_DIR, "cache")
FIXTURE = os.path.join(DATA_DIR, "fixtures", "mcf_sample.json")
ARTIFACT = os.path.join(DATA_DIR, "role_enrichment.json")

# --- MCF API (discovered 2026-07-18: public, unauthenticated GET) ---
MCF_BASE = "https://www.mycareersfuture.gov.sg"
MCF_API = "https://api.mycareersfuture.gov.sg/v2/jobs"
MCF_JOB_URL = "https://www.mycareersfuture.gov.sg/jobs"  # /{uuid}
UA = "AgentProof-research/0.1 (hackathon; +https://github.com/eugenewong22/agentproof)"
CURRENCY = "SGD"
CONFIDENCE_FLOOR = 0.10  # below this, log posting as unmatched (never force-assign)

def _openpyxl():
    import openpyxl, warnings
    warnings.filterwarnings("ignore")
    return openpyxl


@dataclass
class Role:
    sector: str
    track: str
    role: str
    description: str = ""
    key_tasks: list = field(default_factory=list)

    @property
    def key(self) -> str:
        return self.role.strip()

    @property
    def slug(self) -> str:
        s = re.sub(r"[^a-z0-9]+", "-", self.role.lower()).strip("-")
        return s or "role"

    def corpus_text(self) -> str:
        # Text used for TF-IDF matching: role title + description + key tasks.
        return " \n ".join([self.role, self.description, *self.key_tasks])


def load_roles(sector: str, track: str | None, xlsx: str = SFW_XLSX) -> list[Role]:
    """Read SkillsFuture Job Role_Description (+ CWF/KT) filtered by Sector/Track."""
    op = _openpyxl()
    wb = op.load_workbook(xlsx, read_only=True, data_only=True)

    # 1. descriptions
    desc: dict[str, str] = {}
    ws = wb["Job Role_Description"]
    for r in ws.iter_rows(min_row=2, values_only=True):
        if not r[0]:
            continue
        if str(r[0]).strip() != sector:
            continue
        if track and str(r[1]).strip() != track:
            continue
        role_name = str(r[2]).strip()
        desc[role_name] = str(r[3] or "").strip()

    # 2. key tasks (grouped per role)
    kt: dict[str, list[str]] = {r: [] for r in desc}
    ws = wb["Job Role_CWF_KT"]
    for r in ws.iter_rows(min_row=2, values_only=True):
        if not r[0] or str(r[0]).strip() != sector:
            continue
        if track and str(r[1]).strip() != track:
            continue
        role_name = str(r[2]).strip()
        if role_name in kt and r[4]:
            kt[role_name].append(str(r[4]).strip())

    wb.close()
    roles = [
        Role(sector=sector, track=(track or ""), role=name,
             description=desc.get(name, ""), key_tasks=kt.get(name, []))
        for name in desc
    ]
    roles.sort(key=lambda x: x.role)
    return roles


def load_ssoc_map(xlsx: str = SSOC_XLSX) -> dict[str, str]:
    """Best-effort {ssoc_code: title} from the SSOC structure file in this folder.

    MCF postings are tagged with ssocVersion '2020v3'; the repo ships the 2024 file,
    so direct lookups will miss often. SSOC is garnish only (never a hard join),
    so misses are silent and the TF-IDF match carries the decision.
    """
    if not os.path.exists(xlsx):
        return {}
    op = _openpyxl()
    wb = op.load_workbook(xlsx, read_only=True, data_only=True)
    sh = wb.sheetnames[0]
    ws = wb[sh]
    out: dict[str, str] = {}
    for row in ws.iter_rows(min_row=6, values_only=True):  # rows 1-5 are metadata
        code, title = (row[0], row[1]) if len(row) >= 2 else (None, None)
        if code and title:
            out[str(code).strip()] = str(title).strip()
    wb.close()
    return out

class RoleMatcher:
    """TF-IDF + cosine matcher of a posting's text against the SkillsFuture role
    corpus. Same Mode-B design as BUILD_GUIDE Phase 7 (classify.recommend_roles)."""

    def __init__(self, roles: list[Role]):
        from sklearn.feature_extraction.text import TfidfVectorizer
        from sklearn.metrics.pairwise import cosine_similarity
        import numpy as np
        self._np = np
        self.roles = roles
        corpus = [r.corpus_text() for r in roles]
        self._vec = TfidfVectorizer(
            lowercase=True, ngram_range=(1, 2), min_df=1,
            token_pattern=r"(?u)\b[\w./+-]{2,}\b",
        )
        self._M = self._vec.fit_transform(corpus)
        self._cos = cosine_similarity

    def match(self, text: str, ssoc_code: str | None = None,
              ssoc_map: dict[str, str] | None = None) -> tuple[Role | None, float, str]:
        """Return (best_role, confidence, matched_via). confidence in [0,1].

        matched_via: 'tfidf' | 'tfidf+ssoc' | '' (no match above floor).
        SSOC only nudges the reported label; it never overrides the TF-IDF pick.
        """
        if not text.strip():
            return None, 0.0, ""
        v = self._vec.transform([text])
        sims = self._cos(v, self._M)[0]
        i = int(sims.argmax())
        conf = float(sims[i])
        best = self.roles[i]
        via = "tfidf"
        # optional SSOC confirmation
        if ssoc_code and ssoc_map:
            ssoc_title = ssoc_map.get(str(ssoc_code).strip())
            if ssoc_title:
                # does the SSOC title fuzzy-mention the chosen role (or vice versa)?
                a, b = ssoc_title.lower(), best.role.lower()
                if a in b or b in a or _token_overlap(a, b) >= 0.5:
                    via = "tfidf+ssoc"
        if conf < CONFIDENCE_FLOOR:
            return None, conf, ""
        return best, conf, via


def _token_overlap(a: str, b: str) -> float:
    ta, tb = set(a.split()), set(b.split())
    if not ta or not tb:
        return 0.0
    return len(ta & tb) / min(len(ta), len(tb))

# --- fetch (plain requests, no proxy / no browser) ---

def _session() -> requests.Session:
    s = requests.Session()
    s.headers.update({"User-Agent": UA, "Accept": "application/json"})
    return s


def role_query(role: Role) -> str:
    """Best MCF search query for a SkillsFuture role.

    SkillsFuture compound titles like "Data Analyst / Associate Data Engineer"
    over-constrain MCF's keyword search (the slashes make it look for one phrase)
    and return 0 postings. Use the first segment — the primary job title — which
    is what employers actually post under.
    """
    primary = role.role.split("/")[0].strip()
    # drop seniority/level words so we cast a wider net (Senior X, Head of X)
    stop = ("senior", "junior", "associate", "head of", "chief")
    toks = [t for t in primary.split() if t.lower() not in stop]
    return " ".join(toks) or primary or role.role


def endpoint_probe(session: requests.Session, query: str = "data analyst",
                   limit: int = 1) -> int:
    """One live request to confirm the API path/payload. Returns HTTP status."""
    url = MCF_API
    resp = session.get(url, params={"search": query, "limit": limit}, timeout=30)
    print(f"endpoint-probe: GET {url}?search={query}&limit={limit}")
    print(f"  -> HTTP {resp.status_code} | {resp.headers.get('content-type')}")
    if resp.status_code == 200 and resp.headers.get("content-type", "").startswith("application/json"):
        d = resp.json()
        print(f"  top-level keys: {list(d.keys())}")
        print(f"  total={d.get('total')} results={len(d.get('results', []))}")
        if d.get("results"):
            print(f"  result[0] keys: {list(d['results'][0].keys())}")
    else:
        print(f"  body[:200]: {resp.text[:200]}")
    return resp.status_code


def fetch_role(session: requests.Session, role: Role, per_role: int,
               use_cache: bool, delay: float = 2.0) -> dict:
    """Fetch postings for one role. Returns the raw MCF JSON dict.

    Caches per role-slug; --use-cache skips the live request entirely when a cache
    file exists (so demo re-runs cost 0 live requests and are deterministic).
    """
    os.makedirs(CACHE_DIR, exist_ok=True)
    cache_path = os.path.join(CACHE_DIR, f"{role.slug}.json")
    if use_cache and os.path.exists(cache_path):
        with open(cache_path, encoding="utf-8") as f:
            return json.load(f)

    # query = the role's primary title (compound titles over-constrain MCF)
    q = role_query(role)
    params = {"search": q, "limit": per_role, "page": 1}
    for attempt in range(3):
        try:
            resp = session.get(MCF_API, params=params, timeout=30)
            if resp.status_code == 200:
                data = resp.json()
                with open(cache_path, "w", encoding="utf-8") as f:
                    json.dump(data, f, ensure_ascii=False)
                print(f"  [live] {role.role}: {len(data.get('results', []))} postings")
                time.sleep(delay)
                return data
            if resp.status_code in (429, 502, 503):
                wait = delay * (2 ** attempt)
                print(f"  [retry {resp.status_code}] {role.role} — waiting {wait:.0f}s")
                time.sleep(wait)
                continue
            print(f"  [!] {role.role}: HTTP {resp.status_code} — {resp.text[:120]}")
            return {"results": [], "total": 0}
        except requests.RequestException as e:
            wait = delay * (2 ** attempt)
            print(f"  [retry err] {role.role}: {e} — waiting {wait:.0f}s")
            time.sleep(wait)
    print(f"  [!] {role.role}: gave up after retries (cache empty)")
    return {"results": [], "total": 0}


def load_fixture() -> dict:
    """--dry-run source: a small cached MCF JSON shape, 0 live requests."""
    if not os.path.exists(FIXTURE):
        raise FileNotFoundError(
            f"fixture missing: {FIXTURE} — run --endpoint-probe/--live first, "
            f"or drop a sample MCF response there."
        )
    with open(FIXTURE, encoding="utf-8") as f:
        return json.load(f)

# --- posting cleaning + digest ---

_TAG_RE = re.compile(r"<[^>]+>")
_WS_RE = re.compile(r"\s+")


def strip_html(s: str) -> str:
    s = _html.unescape(s or "")
    s = _TAG_RE.sub(" ", s)
    return _WS_RE.sub(" ", s).strip()


def posting_text(job: dict) -> str:
    """Blob for TF-IDF: title + cleaned description + named skills."""
    title = str(job.get("title") or "")
    desc = strip_html(str(job.get("description") or ""))
    skills = ", ".join(s.get("skill", "") for s in (job.get("skills") or []) if s.get("skill"))
    return " \n ".join([title, desc, skills])


def extract_responsibilities(desc_html: str, n: int = 6) -> list[str]:
    """Pull up to N duty lines from a posting's HTML description.

    Looks for explicit list items first; falls back to sentence-ish chunks split
    on sentence terminators / newlines. Purely textual distillation.
    """
    text = strip_html(desc_html)
    if not text:
        return []
    # split on <li> boundaries before stripping would have been nicer; do it on
    # bullet-ish delimiters: newlines, semicolons, or "•"
    parts = re.split(r"[\n;•]|\.(?=\s|$)", text)
    parts = [p.strip(" .:-") for p in parts if len(p.strip()) > 18]
    seen, out = set(), []
    for p in parts:
        key = p.lower()
        if key not in seen:
            seen.add(key)
            out.append(p)
        if len(out) >= n:
            break
    return out


def aggregate_role(role: Role, postings: list[dict], matched_via: str) -> dict:
    """Condense the matched postings for one role into the enrichment entry."""
    from statistics import median
    tools: dict[str, int] = {}
    sal_vals: list[int] = []
    src_urls: list[str] = []
    duties: list[str] = []
    for j in postings:
        uuid = j.get("uuid")
        if uuid:
            src_urls.append(f"{MCF_JOB_URL}/{uuid}")
        for s in (j.get("skills") or []):
            name = (s.get("skill") or "").strip()
            if name:
                tools[name] = tools.get(name, 0) + 1
        sal = j.get("salary") or {}
        if sal.get("minimum") and sal.get("maximum"):
            sal_vals.extend([int(sal["minimum"]), int(sal["maximum"])])
        duties.extend(extract_responsibilities(str(j.get("description") or ""), n=4))

    # keep top tools by frequency, then alpha
    top_tools = [t for t, _ in sorted(tools.items(), key=lambda kv: (-kv[1], kv[0]))[:12]]
    # dedupe duties across postings, keep first n
    seen, resp = set(), []
    for d in duties:
        k = d.lower()
        if k not in seen:
            seen.add(k)
            resp.append(d)
        if len(resp) >= 8:
            break

    band = None
    if sal_vals:
        band = {
            "low": min(sal_vals), "median": int(median(sal_vals)),
            "high": max(sal_vals), "currency": CURRENCY,
        }
    return {
        "sector": role.sector, "track": role.track, "role": role.role,
        "n_postings": len(postings),
        "responsibilities": resp,
        "tools": top_tools,
        "salary_band": band,
        "source_urls": src_urls,
        "matched_via": matched_via,
        "fetched_at": time.strftime("%Y-%m-%dT%H:%M:%S%z", time.gmtime()) + "Z",
    }

# --- orchestration ---

def run(sector: str, track: str | None, per_role: int, use_cache: bool,
        dry_run: bool, probe: bool) -> dict:
    print(f"== AgentProof role-enrichment scraper ==")
    print(f"sector={sector!r} track={track!r} per_role={per_role} "
          f"use_cache={use_cache} dry_run={dry_run} probe={probe}")

    # 1. role list (stand-alone; does NOT import framework.py)
    roles = load_roles(sector, track)
    if not roles:
        print(f"[!] no roles found for sector={sector!r} track={track!r}")
        return {}
    print(f"loaded {len(roles)} SkillsFuture role(s): {[r.role for r in roles]}")

    # 2. fetch (live, cached, or fixture)
    session = _session()
    raw: dict[str, dict] = {}  # role.slug -> MCF json
    if probe:
        endpoint_probe(session)
        return {}
    if dry_run:
        fx = load_fixture()
        # apply the same fixture payload to every target role (proves parse+match only)
        for r in roles:
            raw[r.slug] = fx
        print(f"[dry-run] using fixture {FIXTURE} ({len(fx.get('results', []))} postings) for every role")
    else:
        if use_cache:
            print("[use-cache] skipping live requests where cache exists")
        for r in roles:
            raw[r.slug] = fetch_role(session, r, per_role, use_cache)

    # 3. match + digest
    ssoc_map = load_ssoc_map()
    matcher = RoleMatcher(roles)
    matched: dict[str, dict] = {}     # role.role -> entry
    unmatched: list[str] = []

    for r in roles:
        data = raw.get(r.slug, {})
        postings = data.get("results") or []
        assigned: list[dict] = []
        best_via = ""
        for j in postings:
            role_hit, conf, via = matcher.match(
                posting_text(j), j.get("ssocCode"), ssoc_map
            )
            if role_hit is None:
                unmatched.append(f"{r.role} <- posting '{j.get('title')}' (conf={conf:.2f})")
                continue
            if role_hit.key == r.key:  # posting matched the role we queried for
                assigned.append(j)
                if via and not best_via:
                    best_via = via
            else:
                unmatched.append(
                    f"{r.role} <- posting '{j.get('title')}' matched {role_hit.role} (conf={conf:.2f})"
                )
        if assigned:
            matched[r.key] = aggregate_role(r, assigned, best_via or "tfidf")

    # write artifact
    os.makedirs(DATA_DIR, exist_ok=True)
    with open(ARTIFACT, "w", encoding="utf-8") as f:
        json.dump(matched, f, ensure_ascii=False, indent=2)
    print(f"\nwrote {ARTIFACT} ({len(matched)} role(s) enriched)")
    for name, e in matched.items():
        print(f"  - {name}: {e['n_postings']} postings, "
              f"{len(e['responsibilities'])} duties, {len(e['tools'])} tools, "
              f"via={e['matched_via']}")
    if unmatched:
        print(f"\n{len(unmatched)} unmatched posting(s) logged (confidence floor {CONFIDENCE_FLOOR}):")
        for u in unmatched[:10]:
            print(f"  - {u}")
    return matched


def main(argv: list[str] | None = None) -> int:
    p = argparse.ArgumentParser(description="MyCareersFuture role-enrichment scraper.")
    p.add_argument("--sector", default="Infocomm Technology")
    p.add_argument("--track", default="Data and Artificial Intelligence",
                   help="SkillsFuture track filter (pass '' for whole sector)")
    p.add_argument("--per-role", type=int, default=8, help="postings per role (MCF limit)")
    p.add_argument("--use-cache", action="store_true", help="skip live requests if cache exists")
    p.add_argument("--dry-run", action="store_true", help="fixtures only, 0 live requests")
    p.add_argument("--endpoint-probe", action="store_true", help="one live request, print status, exit")
    a = p.parse_args(argv)
    track = a.track or None
    run(a.sector, track, a.per_role, a.use_cache, a.dry_run, a.endpoint_probe)
    return 0


if __name__ == "__main__":
    sys.exit(main())
